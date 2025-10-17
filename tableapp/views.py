from django.shortcuts import render, redirect, get_object_or_404
from .models import TablePermission, PvoRegistro
from .forms import PvoRegistroForm
from django.contrib.auth.decorators import login_required
from django.db import connections, connection
from django.http import HttpResponse
from django.utils import timezone
from django.utils.timezone import now
from datetime import datetime
from django.shortcuts import render, get_object_or_404
from simple_history.utils import update_change_reason
from two_factor.views import SetupView
from django.shortcuts import redirect
from django_otp.plugins.otp_totp.models import TOTPDevice
from decimal import Decimal
from .models import TablePermission, PvoRegistro, FPConfig, FPCatalogo
from urllib.parse import parse_qs
from io import BytesIO
from django.http import HttpResponse
from io import BytesIO
from decimal import Decimal
from urllib.parse import parse_qs
from datetime import datetime, date
from django.http import HttpResponse
from django.utils.timezone import now
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment   
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, date
from io import BytesIO
from io import BytesIO
from datetime import datetime, date

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db import connections

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

  



def check_perm(user, table, perm):
    try:
        perms = TablePermission.objects.get(user=user, table=table)
        return getattr(perms, f"can_{perm}")
    except TablePermission.DoesNotExist:
        return False


def get_perm_dict(user, table):
    return {
        'can_write': check_perm(user, table, 'write'),
        'can_delete': check_perm(user, table, 'delete'),
        'can_read': check_perm(user, table, 'read'),
        'can_edit': check_perm(user, table, 'edit'),
        'can_edit_full': check_perm(user, table, 'edit_full'),
        'can_edit_flp': check_perm(user, table, 'edit_flp'),
        'can_edit_fef': check_perm(user, table, 'edit_fef'),
        'can_view_history': check_perm(user, table, 'view_history'),
    }


@login_required
def dashboard(request):
    perms = {'report': check_perm(request.user, 'report', 'read')}
    return render(request, 'tableapp/dashboard.html', {'perms': perms})


@login_required
def query_report_view(request):
    if not check_perm(request.user, 'report', 'read'):
        return render(request, 'tableapp/no_permission.html')

    permisos = {
        'can_edit_dates': check_perm(request.user, 'edit_dates', 'edit'),
        'can_edit_full':  check_perm(request.user, 'edit_dates', 'edit_full'),
        'can_edit_flp':   check_perm(request.user, 'edit_dates', 'edit_flp'),
        'can_edit_fif':   check_perm(request.user, 'edit_dates', 'edit_fif'),
        'can_edit_fef':   check_perm(request.user, 'edit_dates', 'edit_fef'),
        'can_view_history': check_perm(request.user, 'report', 'view_history'),
    }

    registros_finales = []

    # 1) Consulta principal (no toco las columnas)
    with connections['ssf_genericos'].cursor() as cursor:
        cursor.execute("""
            SELECT
                CONCAT(in_pedidencab.peeconsecutivo, in_pedidencab.peecompania, in_pediddetal.pedsecuencia) AS PID,
                in_pedidencab.peeconsecutivo AS Pedido,
                in_pedidencab.peeordecompclie AS [OC Cliente],
                in_pedidencab.peecliente AS [Nit Cliente],
                V_SIS_BI_clientesv2.[Razon Social],
                in_pediddetal.pedcodiitem AS [Codigo Producto],
                in_items.itedesclarg AS [Producto Largo],
                in_pedidencab.peefechelab AS [Fecha Pedido],
                in_pediddetal.pedfechrequ AS [Fecha Requerida],
                F.[Fecha Despacho],
                CASE
                    WHEN in_pediddetal.eobnombre IN ('Cerrado', 'Completo') THEN 0
                    ELSE DATEDIFF(DAY, GETDATE(), Op.[Fecha Estimado Fin])
                END AS [Dias de Retraso],
                CASE
                    WHEN in_pediddetal.eobnombre = 'Cerrado' THEN 'Cerrado'
                    WHEN in_pediddetal.eobnombre = 'Completo' THEN 'Despacho Completo'
                    WHEN Op.[Estado Nombre] IN ('En Planeacion', 'En firme', 'Suspendido') THEN 'Compras & ABT'
                    WHEN Op.[Estado Nombre] IN ('Por ejecutar', 'En ejecucion') THEN 'En Produccion'
                    ELSE 'X'
                END AS [Estado Pedido],
                FP.familia          AS [FAMILIA],
                FP.tipo             AS [TIPO],
                FP.tamano_lote      AS [TAMANO_LOTE],
                CASE 
                  WHEN CHARINDEX('X', Prod.[Producto Largo]) > 0 THEN
                    CAST(SUBSTRING(
                      Prod.[Producto Largo],
                      PATINDEX('%[0-9]%', SUBSTRING(Prod.[Producto Largo], CHARINDEX('X', Prod.[Producto Largo]) + 1, LEN(Prod.[Producto Largo])))
                        + CHARINDEX('X', Prod.[Producto Largo]),
                      PATINDEX('%[^0-9]%', SUBSTRING(
                        Prod.[Producto Largo],
                        PATINDEX('%[0-9]%', SUBSTRING(Prod.[Producto Largo], CHARINDEX('X', Prod.[Producto Largo]) + 1, LEN(Prod.[Producto Largo])))
                          + CHARINDEX('X', Prod.[Producto Largo]),
                        LEN(Prod.[Producto Largo])
                      ) + 'X') - 1
                    ) AS INT)
                  ELSE NULL
                END AS [GRAMAJE],
                FP.batch            AS [BATCH],
                FP.estado_fp        AS [ESTADO FP],
                FP.tiempo_fp_horas  AS [TIEMPO FP],
                FP.personal_fase    AS [PERSONAL_FASE],
                FP.planta           AS [PLANTA],
                in_pediddetal.pedcantpediump AS [Cantidad Pedida],
                in_pediddetal.pedcantpediump * in_pediddetal.pedprecunit AS [Valor Pedido],
                in_pediddetal.pedcantdespump AS [Cantidad Despachada],
                in_pediddetal.pedcantdespump * in_pediddetal.pedprecunit AS [Valor Despacho],
                in_pediddetal.pedcantpediump - in_pediddetal.pedcantdespump AS [Cantidad Pendiente],
                (in_pediddetal.pedcantpediump - in_pediddetal.pedcantdespump) * in_pediddetal.pedprecunit AS [Valor Pendiente],
                Op.Op AS OP
            FROM ssf_genericos.dbo.in_pedidencab WITH (NOLOCK)
            INNER JOIN ssf_genericos.dbo.in_pediddetal WITH (NOLOCK)
                ON in_pediddetal.pedconsecutivo = in_pedidencab.peeconsecutivo
               AND in_pediddetal.pedtipocons   = in_pedidencab.peetipocons
               AND in_pediddetal.pedcompania   = in_pedidencab.peecompania
            LEFT OUTER JOIN (
                SELECT
                    pd_ordenproceso.orpcompania AS Compania,
                    MAX(pd_ordenproceso.orpconsecutivo) AS Op,
                    pd_ordenproceso.orpconspedi AS Pedido,
                    pd_ordenproceso.orpsecupedi AS [Secuencia Pedido],
                    pd_ordenproceso.eobcodigo AS Estado,
                    pd_ordenproceso.eobnombre AS [Estado Nombre],
                    pd_ordenproceso.orpfechaentrega AS [Fecha Entrega Planta],
                    pd_ordenproceso.orpfechestifin AS [Fecha Estimado Fin],
                    SUM(pd_ordenproceso.orpcantrecibida) AS [Cantidad Recibida]
                FROM ssf_genericos.dbo.pd_ordenproceso
                WHERE pd_ordenproceso.eobcodigo IN ('PE', 'EP', 'EF', 'EE', 'SU')
                  AND pd_ordenproceso.orpcompania = '01'
                  AND CAST(pd_ordenproceso.orpcantrecibida AS NVARCHAR(15)) + pd_ordenproceso.eobnombre NOT IN ('0.00Cerrado', '0.00Finalizada')
                GROUP BY
                    pd_ordenproceso.orpcompania,
                    pd_ordenproceso.orpconspedi,
                    pd_ordenproceso.orpsecupedi,
                    pd_ordenproceso.eobcodigo,
                    pd_ordenproceso.eobnombre,
                    pd_ordenproceso.orpfechaentrega,
                    pd_ordenproceso.orpfechestifin,
                    pd_ordenproceso.orpcantrecibida
            ) Op
                ON in_pediddetal.pedcompania  = Op.Compania
               AND in_pediddetal.pedconsecutivo = Op.Pedido
               AND in_pediddetal.pedsecuencia  = Op.[Secuencia Pedido]
            LEFT OUTER JOIN (
                SELECT
                    MAX(in_movimientos.movfechmovi) AS [Fecha Despacho],
                    in_movimientos.movconsedocuorig + in_movimientos.movcompania + in_movimientos.movcodiitem AS ID
                FROM ssf_genericos.dbo.in_movimientos
                WHERE in_movimientos.movtipocons IN ('DVTAN', 'DVTAX')
                GROUP BY in_movimientos.movconsedocuorig + in_movimientos.movcompania + in_movimientos.movcodiitem
            ) F
                ON in_pedidencab.peeconsecutivo + in_pedidencab.peecompania + in_pediddetal.pedcodiitem = F.ID
            LEFT OUTER JOIN django_test_db.dbo.tableapp_fpconfig AS FP WITH (NOLOCK)
                ON FP.codigo_producto COLLATE Modern_Spanish_CI_AS = in_pediddetal.pedcodiitem
            LEFT OUTER JOIN ssf_genericos.dbo.V_SIS_BI_Productos AS Prod
                ON Prod.[Codigo Producto] = in_pediddetal.pedcodiitem
               AND Prod.Estado = 'Activo'
            INNER JOIN ssf_genericos.dbo.in_items
                ON in_pediddetal.pedcodiitem  = in_items.itecodigo
               AND in_pediddetal.pedcompania  = in_items.itecompania
            INNER JOIN ssf_genericos.dbo.V_SIS_BI_clientesv2
                ON V_SIS_BI_clientesv2.[Nit Cliente] = in_pedidencab.peecliente
               AND V_SIS_BI_clientesv2.Compañia     = in_pedidencab.peecompania
            LEFT OUTER JOIN django_test_db.dbo.tableapp_pvoregistro
                ON CONCAT(in_pedidencab.peeconsecutivo, in_pedidencab.peecompania, in_pediddetal.pedsecuencia)
                   = tableapp_pvoregistro.pid COLLATE Latin1_General_CI_AS
            WHERE
                YEAR(in_pedidencab.peefechelab) >= YEAR(GETDATE()) - 1
                AND in_pedidencab.peecompania = '01'
                AND in_pediddetal.eobnombre NOT IN ('Cerrado', 'Completo')
                AND in_pedidencab.peetipocons <> 'PECOP'
                AND in_pediddetal.pedcodiitem NOT LIKE '%SER%'
                AND (
                    in_pediddetal.pedrazoncierre IS NULL
                    OR in_pediddetal.pedrazoncierre IN ('07-PEDIDO COMPLETO', '08-PRODUCTO AVERIADO', '03- FACTURADO', '')
                )
            ORDER BY [Fecha Requerida]
        """)
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]

    # 2) Fechas + OBS locales
    with connections['default'].cursor() as cursor2:
        cursor2.execute("""
            SELECT
              pid,
              fecha_full, fecha_flp, fecha_fif, fecha_fef,
              obs_full, obs_flp, obs_fef,
              creado_por_id, fecha_creacion
            FROM tableapp_pvoregistro
        """)
        fechas = cursor2.fetchall()

    fechas_dict = {
        row[0]: {
            'FULL': row[1], 'FLP': row[2], 'FIF': row[3], 'FEF': row[4],
            'OBS_FULL': row[5], 'OBS_FLP': row[6], 'OBS_FEF': row[7],
            'ACTUALIZADO_POR': row[8], 'ACTUALIZADO_EN': row[9],
        } for row in fechas
    }

    # 3) Merge
    for row in rows:
        registro = dict(zip(columns, row))
        pid_base = str(registro['PID']).strip()

        extra = fechas_dict.get(pid_base, {})
        registro['Fecha FULL'] = extra.get('FULL') or registro.get('Fecha FULL')
        registro['Fecha FLP']  = extra.get('FLP')  or registro.get('Fecha FLP')
        registro['Fecha FIF']  = extra.get('FIF')  or registro.get('Fecha FIF')
        registro['Fecha FEF']  = extra.get('FEF')  or registro.get('Fecha FEF')
        registro['Obs FULL'] = extra.get('OBS_FULL')
        registro['Obs FLP']  = extra.get('OBS_FLP')
        registro['Obs FEF']  = extra.get('OBS_FEF')
        registro['Actualizado por'] = extra.get('ACTUALIZADO_POR')
        registro['Última Fecha']    = extra.get('ACTUALIZADO_EN')

        # Sincroniza capacidad con GRAMAJE (no crítico si falla)
        try:
            codigo = registro.get('Codigo Producto')
            gramaje = registro.get('GRAMAJE')
            if codigo and gramaje is not None:
                fp, _ = FPConfig.objects.get_or_create(codigo_producto=codigo)
                if fp.capacidad != gramaje:
                    fp.capacidad = gramaje
                    fp.save(update_fields=['capacidad'])
        except Exception:
            pass

        registros_finales.append(registro)

    skip_cols = ['Fecha Entrega Planta', 'Fecha Estimado Fin',
                 'Fecha FULL', 'Fecha FLP', 'Fecha FIF', 'Fecha FEF']

    # 4) Catálogo
    orden_fp = ['DISPENSACION','PESAJE','FABRICACION','ENFRIAMIENTO','ENVASADO',
                'ACONDICIONAMIENTO','EMBALAJE','DESPACHO']
    vals_estados = list(FPCatalogo.objects.filter(grupo='ESTADO FP')
                        .values_list('valor', flat=True))
    def _k(v):
        u = (v or '').strip().upper()
        return (orden_fp.index(u) if u in orden_fp else len(orden_fp), u)
    estados_ordenados = sorted(vals_estados, key=_k)

    catalogo = {
        'estados_fp': estados_ordenados,
        'familias': list(FPCatalogo.objects.filter(grupo='FAMILIA')
                         .values_list('valor', flat=True).order_by('valor')),
        'tipos': list(FPCatalogo.objects.filter(grupo='TIPO')
                      .values_list('valor', flat=True).order_by('valor')),
        'plantas': list(FPCatalogo.objects.filter(grupo='PLANTA')
                        .values_list('valor', flat=True).order_by('valor')),
    }

    return render(request, 'tableapp/query_report.html', {
        'registros': registros_finales,
        'columns': columns,
        'perms': permisos,
        'skip_cols': skip_cols,
        'catalogo': catalogo,
    })


@login_required
def historial_pvo_view(request):
    if not check_perm(request.user, 'report', 'read'):
        return render(request, 'tableapp/no_permission.html')

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT * FROM (
                SELECT pid AS guid, 'Compras' AS [Área], fecha_full AS [Fecha],
                       creado_por_id AS actualizado_por, fecha_creacion AS actualizado_en, '' AS comentario
                FROM tableapp_pvoregistro WHERE fecha_full IS NOT NULL
                UNION ALL
                SELECT pid, 'Liberación', fecha_flp, creado_por_id, fecha_creacion, ''
                FROM tableapp_pvoregistro WHERE fecha_flp IS NOT NULL
                UNION ALL
                SELECT pid, 'Inicio Fab.', fecha_fif, creado_por_id, fecha_creacion, ''
                FROM tableapp_pvoregistro WHERE fecha_fif IS NOT NULL
                UNION ALL
                SELECT pid, 'Producción', fecha_fef, creado_por_id, fecha_creacion, ''
                FROM tableapp_pvoregistro WHERE fecha_fef IS NOT NULL
            ) AS historial
            ORDER BY guid, actualizado_en DESC
        """)
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]

    return render(request, 'tableapp/query_results.html', {'columns': columns, 'rows': rows})


@login_required
def pvo_list(request):
    if not check_perm(request.user, 'report', 'read'):
        return render(request, 'tableapp/no_permission.html')
    registros = PvoRegistro.objects.all()
    perms = get_perm_dict(request.user, 'report')
    return render(request, 'tableapp/pvo_list.html', {'registros': registros, 'perms': perms})


@login_required
def pvo_create(request):
    if not check_perm(request.user, 'report', 'write'):
        return render(request, 'tableapp/no_permission.html')

    form = PvoRegistroForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        registro = form.save(commit=False)
        registro.actualizado_por = request.user
        registro.actualizado_en = timezone.now()
        registro.save()
        if request.htmx:
            return HttpResponse(status=204)
        return redirect('pvo_list')

    perms = get_perm_dict(request.user, 'report')
    return render(request, 'tableapp/form.html', {'form': form, 'cancel_url': 'pvo_list', 'perms': perms})


@login_required
def pvo_edit(request, pk):
    if not check_perm(request.user, 'report', 'edit'):
        return render(request, 'tableapp/no_permission.html')

    registro = get_object_or_404(PvoRegistro, pk=pk)
    form = PvoRegistroForm(request.POST or None, instance=registro)

    if request.method == 'POST' and form.is_valid():
        registro = form.save(commit=False)
        registro.actualizado_por = request.user
        registro.actualizado_en = timezone.now()
        registro.save()
        if request.htmx:
            return HttpResponse(status=204)
        return redirect('pvo_list')

    perms = get_perm_dict(request.user, 'report')
    return render(request, 'tableapp/form.html', {'form': form, 'cancel_url': 'pvo_list', 'perms': perms, 'object': registro})


@login_required
def actualizar_fecha(request, pid, campo):
    """
    Actualiza:
      - FULL / FLP / FIF / FEF con input name="fecha" (YYYY-MM-DD)
      - OBS_FULL / OBS_FLP / OBS_FEF con input name="obs"
    """
    if request.method != 'PUT':
        return HttpResponse(status=405)

    try:
        data = {k: v[0] for k, v in parse_qs(request.body.decode('utf-8')).items()}
        fecha_nueva = (data.get('fecha') or '').strip()
        obs_nueva   = (data.get('obs') or '').strip()

        registro, _ = PvoRegistro.objects.get_or_create(pid=pid)
        update_change_reason(registro, f"{campo} actualizado por {request.user.username}")
        registro._history_user = request.user

        def to_date(s):
            return datetime.strptime(s, '%Y-%m-%d') if s else None

        if campo == 'FULL':
            registro.fecha_full = to_date(fecha_nueva)
        elif campo == 'FLP':
            registro.fecha_flp = to_date(fecha_nueva)
        elif campo == 'FIF':
            registro.fecha_fif = to_date(fecha_nueva)
        elif campo == 'FEF':
            registro.fecha_fef = to_date(fecha_nueva)
        elif campo == 'OBS_FULL':
            registro.obs_full = obs_nueva or None
        elif campo == 'OBS_FLP':
            registro.obs_flp = obs_nueva or None
        elif campo == 'OBS_FEF':
            registro.obs_fef = obs_nueva or None
        else:
            return HttpResponse('Campo no permitido', status=400)

        registro.creado_por = request.user
        registro.fecha_creacion = now()
        registro.save()
        return HttpResponse(status=204)

    except Exception as e:
        print(f"Error actualizando {pid}: {e}")
        return HttpResponse(f"Error: {e}", status=400)


@login_required
def pvo_historial_modal(request, pid):
    # ¡IMPORTANTE! Esta vista existe para que no falle la URL.
    if not check_perm(request.user, 'edit_dates', 'view_history'):
        return render(request, 'tableapp/no_permission.html')

    registro = get_object_or_404(PvoRegistro, pid=pid)
    historico = []

    campos_visibles = [
        field.name for field in registro.history.model._meta.fields
        if field.name not in (
            'id', 'history_id', 'history_date', 'history_user',
            'history_type', 'history_change_reason',
            'creado_por', 'fecha_creacion'
        )
    ]

    for h in registro.history.all().order_by('-history_date'):
        cambios_reales = []
        if h.prev_record:
            for field in campos_visibles:
                old = getattr(h.prev_record, field, None)
                new = getattr(h, field, None)
                if old != new:
                    cambios_reales.append({'field': field, 'old': old, 'new': new})

        if cambios_reales:
            historico.append({'history': h, 'cambios': cambios_reales})

    return render(request, 'tableapp/pvo_historial_modal.html', {'registro': registro, 'historico': historico})


@login_required
def login_redirect_view(request):
    user = request.user
    if not TOTPDevice.objects.filter(user=user, confirmed=True).exists():
        return redirect('/account/two_factor/setup/')
    return redirect('dashboard')


class CustomSetupView(SetupView):
    def done(self, form_list, **kwargs):
        return redirect('dashboard')


@login_required
def actualizar_fp(request, codigo_producto, campo):
    if request.method != 'PUT':
        return HttpResponse(status=405)
    try:
        body = request.body.decode('utf-8')
        parts = [p for p in body.split('&') if '=' in p]
        data = dict(p.split('=', 1) for p in parts)

        valor = (data.get('valor') or '').strip()
        pid   = (data.get('pid') or '').strip()

        permitidos = {'estado_fp', 'tiempo_fp_horas', 'familia', 'tipo',
                      'batch', 'planta', 'tamano_lote', 'personal_fase'}
        if campo not in permitidos:
            return HttpResponse('Campo no permitido', status=400)

        if valor == '':
            nuevo_valor = None
        elif campo == 'tiempo_fp_horas':
            try:
                nuevo_valor = Decimal(valor.replace(',', '.'))
            except Exception:
                return HttpResponse('Valor inválido para TIEMPO FP (horas)', status=400)
        elif campo == 'personal_fase':
            try:
                nuevo_valor = int(valor)
            except Exception:
                return HttpResponse('Valor inválido para PERSONAL_FASE', status=400)
        else:
            nuevo_valor = valor

        fp, _ = FPConfig.objects.get_or_create(codigo_producto=codigo_producto)
        setattr(fp, campo, nuevo_valor)
        fp.save()

        if pid:
            reg, _ = PvoRegistro.objects.get_or_create(pid=pid)

            fase_map = {
                'DISPENSACION': ('f_dispensacion', 't_dispensacion', 'p_dispensacion'),
                'PESAJE': ('f_pesaje', 't_pesaje', 'p_pesaje'),
                'FABRICACION': ('f_fabricacion', 't_fabricacion', 'p_fabricacion'),
                'ENFRIAMIENTO': ('f_enfriamiento', 't_enfriamiento', 'p_enfriamiento'),
                'ENVASADO': ('f_envasado', 't_envasado', 'p_envasado'),
                'ACONDICIONAMIENTO': ('f_acondicionamiento', 't_acondicionamiento', 'p_acondicionamiento'),
                'EMBALAJE': ('f_embalaje', 't_embalaje', 'p_embalaje'),
                'DESPACHO': ('f_despacho', 't_despacho', 'p_despacho'),
            }

            if campo in {'batch', 'planta', 'tamano_lote', 'familia', 'tipo'}:
                setattr(reg, campo, nuevo_valor)

            def parece_fecha(v):
                if v is None:
                    return False
                s = str(v)
                return len(s) >= 10 and s[4] == '-' and s[7] == '-'

            if campo == 'estado_fp':
                fase = (valor or '').upper()
                cols = fase_map.get(fase)
                if cols:
                    col_f, _, _ = cols
                    setattr(reg, col_f, fase)

            if campo in {'tiempo_fp_horas', 'personal_fase'}:
                fase_actual = (fp.estado_fp or '').upper()
                cols = fase_map.get(fase_actual)
                if not cols:
                    return HttpResponse('Define primero ESTADO FP para esta fila.', status=400)
                col_f, col_t, col_p = cols

                f_val = getattr(reg, col_f, None)
                if not f_val or parece_fecha(f_val):
                    setattr(reg, col_f, fase_actual)

                if campo == 'tiempo_fp_horas':
                    setattr(reg, col_t, nuevo_valor)
                else:
                    setattr(reg, col_p, nuevo_valor)

            reg.creado_por = request.user
            reg.fecha_creacion = now()
            reg.save()

        return HttpResponse(status=204)

    except Exception as e:
        print(f'Error actualizar_fp [{codigo_producto}::{campo}]: {e}')
        return HttpResponse(f'Error: {e}', status=400)


def _match_filters(registro: dict, filtros: dict) -> bool:
    """Devuelve True si el dict 'registro' cumple todos los filtros 'contains'."""
    for k, v in filtros.items():
        if not v:
            continue
        val = registro.get(k, "")
        if val is None:
            return False
        if isinstance(val, (int, float)):
            val = str(val)
        elif isinstance(val, (datetime, date)):
            val = val.strftime("%Y-%m-%d")
        if v.casefold() not in str(val).casefold():
            return False
    return True




# ... (tus otros imports y funciones siguen igual: check_perm, etc.)

@login_required
def query_report_export(request):
    # Permiso
    if not check_perm(request.user, 'report', 'read'):
        return render(request, 'tableapp/no_permission.html')

    # ---- 0) ¿Exportar solo PIDs visibles? ----
    pids_param = (request.GET.get('pids') or '').strip()
    pids = [p.strip() for p in pids_param.split(',') if p.strip()] if pids_param else []

    registros_finales = []

    # ---- 1) SQL base (misma consulta) ----
    with connections['ssf_genericos'].cursor() as cursor:
        base_sql = """
            SELECT
                CONCAT(in_pedidencab.peeconsecutivo, in_pedidencab.peecompania, in_pediddetal.pedsecuencia) AS PID,
                in_pedidencab.peeconsecutivo AS Pedido,
                in_pedidencab.peeordecompclie AS [OC Cliente],
                in_pedidencab.peecliente AS [Nit Cliente],
                V_SIS_BI_clientesv2.[Razon Social],
                in_pediddetal.pedcodiitem AS [Codigo Producto],
                in_items.itedesclarg AS [Producto Largo],
                in_pedidencab.peefechelab AS [Fecha Pedido],
                in_pediddetal.pedfechrequ AS [Fecha Requerida],
                F.[Fecha Despacho],
                CASE
                    WHEN in_pediddetal.eobnombre IN ('Cerrado', 'Completo') THEN 0
                    ELSE DATEDIFF(DAY, GETDATE(), Op.[Fecha Estimado Fin])
                END AS [Dias de Retraso],
                CASE
                    WHEN in_pediddetal.eobnombre = 'Cerrado' THEN 'Cerrado'
                    WHEN in_pediddetal.eobnombre = 'Completo' THEN 'Despacho Completo'
                    WHEN Op.[Estado Nombre] IN ('En Planeacion', 'En firme', 'Suspendido') THEN 'Compras & ABT'
                    WHEN Op.[Estado Nombre] IN ('Por ejecutar', 'En ejecucion') THEN 'En Produccion'
                    ELSE 'X'
                END AS [Estado Pedido],
                FP.familia          AS [FAMILIA],
                FP.tipo             AS [TIPO],
                FP.tamano_lote      AS [TAMANO_LOTE],
                CASE 
                  WHEN CHARINDEX('X', Prod.[Producto Largo]) > 0 THEN
                    CAST(SUBSTRING(
                      Prod.[Producto Largo],
                      PATINDEX('%[0-9]%', SUBSTRING(Prod.[Producto Largo], CHARINDEX('X', Prod.[Producto Largo]) + 1, LEN(Prod.[Producto Largo])))
                        + CHARINDEX('X', Prod.[Producto Largo]),
                      PATINDEX('%[^0-9]%', SUBSTRING(
                        Prod.[Producto Largo],
                        PATINDEX('%[0-9]%', SUBSTRING(Prod.[Producto Largo], CHARINDEX('X', Prod.[Producto Largo]) + 1, LEN(Prod.[Producto Largo])))
                          + CHARINDEX('X', Prod.[Producto Largo]),
                        LEN(Prod.[Producto Largo])
                      ) + 'X') - 1
                    ) AS INT)
                  ELSE NULL
                END AS [GRAMAJE],
                FP.batch            AS [BATCH],
                FP.estado_fp        AS [ESTADO FP],
                FP.tiempo_fp_horas  AS [TIEMPO FP],
                FP.personal_fase    AS [PERSONAL_FASE],
                FP.planta           AS [PLANTA],
                in_pediddetal.pedcantpediump AS [Cantidad Pedida],
                in_pediddetal.pedcantpediump * in_pediddetal.pedprecunit AS [Valor Pedido],
                in_pediddetal.pedcantdespump AS [Cantidad Despachada],
                in_pediddetal.pedcantdespump * in_pediddetal.pedprecunit AS [Valor Despacho],
                in_pediddetal.pedcantpediump - in_pediddetal.pedcantdespump AS [Cantidad Pendiente],
                (in_pediddetal.pedcantpediump - in_pediddetal.pedcantdespump) * in_pediddetal.pedprecunit AS [Valor Pendiente],
                Op.Op AS OP
            FROM ssf_genericos.dbo.in_pedidencab WITH (NOLOCK)
            INNER JOIN ssf_genericos.dbo.in_pediddetal WITH (NOLOCK)
                ON in_pediddetal.pedconsecutivo = in_pedidencab.peeconsecutivo
               AND in_pediddetal.pedtipocons   = in_pedidencab.peetipocons
               AND in_pediddetal.pedcompania   = in_pedidencab.peecompania
            LEFT OUTER JOIN (
                SELECT
                    pd_ordenproceso.orpcompania AS Compania,
                    MAX(pd_ordenproceso.orpconsecutivo) AS Op,
                    pd_ordenproceso.orpconspedi AS Pedido,
                    pd_ordenproceso.orpsecupedi AS [Secuencia Pedido],
                    pd_ordenproceso.eobcodigo AS Estado,
                    pd_ordenproceso.eobnombre AS [Estado Nombre],
                    pd_ordenproceso.orpfechaentrega AS [Fecha Entrega Planta],
                    pd_ordenproceso.orpfechestifin AS [Fecha Estimado Fin],
                    SUM(pd_ordenproceso.orpcantrecibida) AS [Cantidad Recibida]
                FROM ssf_genericos.dbo.pd_ordenproceso
                WHERE pd_ordenproceso.eobcodigo IN ('PE', 'EP', 'EF', 'EE', 'SU')
                  AND pd_ordenproceso.orpcompania = '01'
                  AND CAST(pd_ordenproceso.orpcantrecibida AS NVARCHAR(15)) + pd_ordenproceso.eobnombre NOT IN ('0.00Cerrado', '0.00Finalizada')
                GROUP BY
                    pd_ordenproceso.orpcompania,
                    pd_ordenproceso.orpconspedi,
                    pd_ordenproceso.orpsecupedi,
                    pd_ordenproceso.eobcodigo,
                    pd_ordenproceso.eobnombre,
                    pd_ordenproceso.orpfechaentrega,
                    pd_ordenproceso.orpfechestifin,
                    pd_ordenproceso.orpcantrecibida
            ) Op
                ON in_pediddetal.pedcompania  = Op.Compania
               AND in_pediddetal.pedconsecutivo = Op.Pedido
               AND in_pediddetal.pedsecuencia  = Op.[Secuencia Pedido]
            LEFT OUTER JOIN (
                SELECT
                    MAX(in_movimientos.movfechmovi) AS [Fecha Despacho],
                    in_movimientos.movconsedocuorig + in_movimientos.movcompania + in_movimientos.movcodiitem AS ID
                FROM ssf_genericos.dbo.in_movimientos
                WHERE in_movimientos.movtipocons IN ('DVTAN', 'DVTAX')
                GROUP BY in_movimientos.movconsedocuorig + in_movimientos.movcompania + in_movimientos.movcodiitem
            ) F
                ON in_pedidencab.peeconsecutivo + in_pedidencab.peecompania + in_pediddetal.pedcodiitem = F.ID
            LEFT OUTER JOIN django_test_db.dbo.tableapp_fpconfig AS FP WITH (NOLOCK)
                ON FP.codigo_producto COLLATE Modern_Spanish_CI_AS = in_pediddetal.pedcodiitem
            LEFT OUTER JOIN ssf_genericos.dbo.V_SIS_BI_Productos AS Prod
                ON Prod.[Codigo Producto] = in_pediddetal.pedcodiitem
               AND Prod.Estado = 'Activo'
            INNER JOIN ssf_genericos.dbo.in_items
                ON in_pediddetal.pedcodiitem  = in_items.itecodigo
               AND in_pediddetal.pedcompania  = in_items.itecompania
            INNER JOIN ssf_genericos.dbo.V_SIS_BI_clientesv2
                ON V_SIS_BI_clientesv2.[Nit Cliente] = in_pedidencab.peecliente
               AND V_SIS_BI_clientesv2.Compañia     = in_pedidencab.peecompania
            LEFT OUTER JOIN django_test_db.dbo.tableapp_pvoregistro
                ON CONCAT(in_pedidencab.peeconsecutivo, in_pedidencab.peecompania, in_pediddetal.pedsecuencia)
                   = tableapp_pvoregistro.pid COLLATE Latin1_General_CI_AS
            WHERE
                YEAR(in_pedidencab.peefechelab) >= YEAR(GETDATE()) - 1
                AND in_pedidencab.peecompania = '01'
                AND in_pediddetal.eobnombre NOT IN ('Cerrado', 'Completo')
                AND in_pedidencab.peetipocons <> 'PECOP'
                AND in_pediddetal.pedcodiitem NOT LIKE '%SER%'
                AND (
                    in_pediddetal.pedrazoncierre IS NULL
                    OR in_pediddetal.pedrazoncierre IN ('07-PEDIDO COMPLETO', '08-PRODUCTO AVERIADO', '03- FACTURADO', '')
                )
        """

        # ---- 1.1) Filtro por PIDs sin usar parámetros (evita choque con % del SQL) ----
        if pids:
            seguros = [("'" + p.replace("'", "''") + "'") for p in pids]
            in_clause = ",".join(seguros)
            base_sql += f"""
                AND CONCAT(in_pedidencab.peeconsecutivo, in_pedidencab.peecompania, in_pediddetal.pedsecuencia) IN ({in_clause})
            """

        base_sql += " ORDER BY [Fecha Requerida]"

        # Ejecutamos SIN params para no disparar el formateo debug con %.
        cursor.execute(base_sql)
        rows = cursor.fetchall()
        columns = [c[0] for c in cursor.description]

    # ---- 2) Fechas/obs locales ----
    with connections['default'].cursor() as c2:
        c2.execute("""
            SELECT pid, fecha_full, fecha_flp, fecha_fif, fecha_fef,
                   obs_full, obs_flp, obs_fef, creado_por_id, fecha_creacion
            FROM tableapp_pvoregistro
        """)
        fechas = c2.fetchall()

    fechas_dict = {
        r[0]: {
            'FULL': r[1], 'FLP': r[2], 'FIF': r[3], 'FEF': r[4],
            'OBS_FULL': r[5], 'OBS_FLP': r[6], 'OBS_FEF': r[7],
            'ACTUALIZADO_POR': r[8], 'ACTUALIZADO_EN': r[9],
        } for r in fechas
    }

    registros_finales = []
    for r in rows:
        d = dict(zip(columns, r))
        pid = str(d['PID']).strip()
        extra = fechas_dict.get(pid, {})
        d['Fecha FULL'] = extra.get('FULL') or d.get('Fecha FULL')
        d['Fecha FLP']  = extra.get('FLP')  or d.get('Fecha FLP')
        d['Fecha FIF']  = extra.get('FIF')  or d.get('Fecha FIF')
        d['Fecha FEF']  = extra.get('FEF')  or d.get('Fecha FEF')
        d['Obs FULL']   = extra.get('OBS_FULL')
        d['Obs FLP']    = extra.get('OBS_FLP')
        d['Obs FEF']    = extra.get('OBS_FEF')
        d['Actualizado por'] = extra.get('ACTUALIZADO_POR')
        d['Última Fecha']    = extra.get('ACTUALIZADO_EN')
        registros_finales.append(d)

    # ---- 3) Excel con estilos ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte PVO"

    columnas_export = columns + [
        'Fecha FULL', 'Obs FULL',
        'Fecha FLP', 'Obs FLP',
        'Fecha FIF',
        'Fecha FEF', 'Obs FEF',
        'Actualizado por', 'Última Fecha'
    ]

    # Encabezados
    ws.append(columnas_export)
    header_fill = PatternFill("solid", fgColor="1e293b")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, len(columnas_export) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border_all

    # Datos
    for item in registros_finales:
        ws.append([item.get(col) for col in columnas_export])

    # Formato fechas
    nombres_fecha = {'Fecha Pedido','Fecha Requerida','Fecha Despacho',
                     'Fecha FULL','Fecha FLP','Fecha FIF','Fecha FEF','Última Fecha'}
    idx_fechas = [i+1 for i, c in enumerate(columnas_export) if c in nombres_fecha]

    for r in range(2, ws.max_row + 1):
        # zebra
        if r % 2 == 0:
            for c in range(1, len(columnas_export)+1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="F7FAFC")
        # borders + wrap para texto largo
        for c in range(1, len(columnas_export)+1):
            cell = ws.cell(row=r, column=c)
            cell.border = border_all
            if columnas_export[c-1] in ('Producto Largo','Razon Social','Obs FULL','Obs FLP','Obs FEF'):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        # fechas
        for c in idx_fechas:
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (datetime, date)) or (isinstance(cell.value, str) and cell.value.count('-') == 2):
                cell.number_format = "DD/MM/YYYY"
                cell.alignment = center

    # Anchos automáticos (tope 45)
    for i, name in enumerate(columnas_export, start=1):
        max_len = max(
            len(str(name)),
            *[len(str(ws.cell(row=r, column=i).value or "")) for r in range(2, ws.max_row+1)]
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 45)

    # Congelar encabezado y AutoFilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Descargar
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Reporte_PVO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
